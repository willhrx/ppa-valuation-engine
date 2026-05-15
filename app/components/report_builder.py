"""PDF and Excel report builders.

Both functions return raw bytes so the Streamlit ``download_button`` can
serve them directly.
"""

from __future__ import annotations

import io
from dataclasses import asdict
from datetime import datetime
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from fpdf import FPDF

from app.components import charts
from ppa_engine.config import PPAConfig


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _fig_to_png(fig: go.Figure) -> bytes:
    """Render a Plotly figure to PNG via Kaleido. Falls back to empty bytes."""
    try:
        return fig.to_image(format="png", scale=2)
    except Exception:
        return b""


class _PDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(120, 120, 120)
        self.cell(0, 8, "PPA Valuation Report", align="L")
        self.cell(0, 8, datetime.utcnow().strftime("Generated %Y-%m-%d %H:%M UTC"), align="R")
        self.ln(12)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 8, f"Page {self.page_no()}", align="C")


def _h(pdf: _PDF, text: str, size: int = 14) -> None:
    pdf.set_font("Helvetica", "B", size)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, text, ln=1)


def _p(pdf: _PDF, text: str, size: int = 10) -> None:
    pdf.set_font("Helvetica", "", size)
    pdf.set_text_color(30, 41, 59)
    pdf.multi_cell(0, 5, text)


def _add_image(pdf: _PDF, png_bytes: bytes, *, w: float = 180, h: float | None = None) -> None:
    if not png_bytes:
        _p(pdf, "[chart could not be rendered — install kaleido]")
        return
    bio = io.BytesIO(png_bytes)
    kwargs: dict[str, Any] = {"w": w}
    if h is not None:
        kwargs["h"] = h
    pdf.image(bio, **kwargs)
    pdf.ln(2)


def _table(pdf: _PDF, df: pd.DataFrame, col_widths: list[float] | None = None) -> None:
    if df.empty:
        return
    if col_widths is None:
        total = 190.0
        col_widths = [total / len(df.columns)] * len(df.columns)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(241, 245, 249)
    for col, w in zip(df.columns, col_widths):
        pdf.cell(w, 7, str(col)[:30], border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(30, 41, 59)
    for _, row in df.iterrows():
        for val, w in zip(row.values, col_widths):
            if isinstance(val, float):
                txt = f"{val:,.2f}"
            else:
                txt = str(val)
            pdf.cell(w, 6, txt[:30], border=1)
        pdf.ln()


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------


def build_pdf_report(
    config: PPAConfig,
    mc_result,
    risk_summary,
    solver_results: dict,
) -> bytes:
    pdf = _PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    # ---- Page 1 — Cover ------------------------------------------------
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 22)
    pdf.set_text_color(15, 23, 42)
    pdf.ln(40)
    pdf.cell(0, 14, "PPA Valuation Report", align="C", ln=1)

    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 8, "Scenario-conditional valuation and risk decomposition", align="C", ln=1)
    pdf.ln(10)

    deal_rows = pd.DataFrame(
        [
            ("Asset capacity", f"{config.solar.capacity_mw:.1f} MW"),
            ("Location", f"{config.location.lat:.2f}°N, {config.location.lon:.2f}°E"),
            ("PPA tenor", f"{config.deal.start_date} → {config.deal.end_date}"),
            ("Offtaker", f"{config.load.annual_consumption_gwh:.1f} GWh / yr"),
            ("Discount rate", f"{config.deal.discount_rate*100:.2f} %"),
            ("Base strike", f"€{mc_result.base_strike:.2f} / MWh"),
        ],
        columns=["Parameter", "Value"],
    )
    _table(pdf, deal_rows, col_widths=[80, 110])
    pdf.ln(8)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, datetime.utcnow().strftime("Generated %Y-%m-%d %H:%M UTC"), align="L", ln=1)

    # ---- Page 2 — Executive summary -----------------------------------
    pdf.add_page()
    _h(pdf, "Executive summary")

    pap_ff = next(
        c for c in risk_summary.combinations
        if c.supply_structure == "PayAsProduced" and c.pricing_structure == "FixedFlat"
    )
    central_row = mc_result.central_df[
        (mc_result.central_df.supply_structure == "PayAsProduced")
        & (mc_result.central_df.pricing_structure == "FixedFlat")
    ].iloc[0]

    pap_neg = solver_results.get("negotiation", {}).get("PayAsProduced")
    floor_str = f"€{pap_neg.producer_floor.strike_eur_mwh:.2f}/MWh" if pap_neg else "—"
    spread_str = f"€{pap_neg.spread_eur_mwh:.2f}/MWh" if pap_neg else "—"

    summary_df = pd.DataFrame(
        [
            ("Central producer NPV", f"€{central_row.producer_npv/1e6:,.2f} M"),
            ("P50 NPV (joint)", f"€{pap_ff.joint.p50/1e6:,.2f} M"),
            ("P10–P90 range", f"€{pap_ff.joint.p10/1e6:,.2f} M to €{pap_ff.joint.p90/1e6:,.2f} M"),
            ("Break-even strike", floor_str),
            ("Negotiation spread", spread_str),
            ("Central capture rate", f"{pap_ff.central_capture_rate*100:.1f} %"),
        ],
        columns=["Metric", "Value"],
    )
    _table(pdf, summary_df, col_widths=[80, 110])
    pdf.ln(4)
    _p(
        pdf,
        f"For a PayAsProduced / FixedFlat structure at €{mc_result.base_strike:.1f}/MWh, "
        f"the P50 producer NPV is €{pap_ff.joint.p50/1e6:.2f} M with a P10–P90 "
        f"range of €{(pap_ff.joint.p90 - pap_ff.joint.p10)/1e6:.2f} M. "
        f"The probability of a negative producer NPV is {pap_ff.joint.prob_negative*100:.1f}%.",
    )

    # ---- Page 3 — NPV distribution & risk table ----------------------
    pdf.add_page()
    _h(pdf, "NPV distribution — PayAsProduced / FixedFlat")

    paths_df = mc_result.paths_df
    mask = (
        (paths_df.supply_structure == "PayAsProduced")
        & (paths_df.pricing_structure == "FixedFlat")
    )
    raw = {
        "joint": paths_df[mask & (paths_df["mode"] == "joint")]["producer_npv"].to_numpy(),
        "price": paths_df[mask & (paths_df["mode"] == "price")]["producer_npv"].to_numpy(),
        "volume": paths_df[mask & (paths_df["mode"] == "volume")]["producer_npv"].to_numpy(),
    }
    pap_ff.raw = raw  # type: ignore[attr-defined]
    _add_image(pdf, _fig_to_png(charts.chart_npv_distribution(pap_ff)))

    rows = []
    for c in risk_summary.combinations:
        rows.append([
            c.supply_structure, c.pricing_structure,
            c.joint.p10 / 1e6, c.joint.p50 / 1e6, c.joint.p90 / 1e6,
            c.joint.var_95 / 1e6, c.joint.es_95 / 1e6,
            c.joint.prob_negative * 100,
        ])
    risk_df = pd.DataFrame(rows, columns=[
        "Supply", "Pricing", "P10", "P50", "P90", "VaR-95", "ES-95", "P<0%",
    ])
    _table(pdf, risk_df)

    # ---- Page 4 — Break-even and tornado ------------------------------
    pdf.add_page()
    _h(pdf, "Break-even and negotiation range")
    _add_image(pdf, _fig_to_png(charts.chart_negotiation_range(solver_results)))

    if "negotiation" in solver_results:
        neg_rows = []
        for supply_name, neg in solver_results["negotiation"].items():
            neg_rows.append([
                supply_name,
                neg.producer_floor.strike_eur_mwh,
                neg.consumer_ceiling.strike_eur_mwh,
                neg.midpoint,
                neg.spread_eur_mwh,
            ])
        _table(
            pdf,
            pd.DataFrame(neg_rows, columns=["Supply", "Floor", "Ceiling", "Midpoint", "Spread"]),
        )

    tornado = solver_results.get("tornado_pap_fixedflat")
    if tornado is not None:
        pdf.ln(4)
        _h(pdf, "Tornado — NPV sensitivity", size=12)
        _add_image(pdf, _fig_to_png(charts.chart_tornado(tornado)))

    # ---- Page 5 — Risk attribution -----------------------------------
    pdf.add_page()
    _h(pdf, "Risk attribution")
    _add_image(pdf, _fig_to_png(charts.chart_variance_decomposition(risk_summary)))
    _p(
        pdf,
        "Three-ensemble variance decomposition: total NPV variance is "
        "attributed to price risk (price-only ensemble), volume risk "
        "(volume-only ensemble), and the residual interaction term. A "
        "negative interaction term indicates diversification or anti-"
        "correlation between price and volume — characteristic of solar "
        "cannibalisation under PayAsProduced contracts.",
    )

    # ---- Page 6 — Assumptions register -------------------------------
    pdf.add_page()
    _h(pdf, "Assumptions register")

    def _flatten(prefix: str, obj: Any, rows: list[tuple[str, str]]):
        for k, v in asdict(obj).items() if hasattr(obj, "__dataclass_fields__") else obj.items():
            name = f"{prefix}.{k}" if prefix else k
            if isinstance(v, list):
                rows.append((name, f"len={len(v)}"))
            else:
                rows.append((name, str(v)))

    rows_assumptions: list[tuple[str, str]] = []
    cfg_dict = asdict(config)
    for section, vals in cfg_dict.items():
        if isinstance(vals, dict):
            for k, v in vals.items():
                if isinstance(v, list):
                    rows_assumptions.append((f"{section}.{k}", f"len={len(v)} list"))
                else:
                    rows_assumptions.append((f"{section}.{k}", str(v)))

    _table(
        pdf,
        pd.DataFrame(rows_assumptions, columns=["Parameter", "Value"]),
        col_widths=[80, 110],
    )

    raw_bytes = pdf.output(dest="S")
    if isinstance(raw_bytes, str):
        return raw_bytes.encode("latin-1")
    return bytes(raw_bytes)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------


def build_excel_report(
    config: PPAConfig,
    mc_result,
    risk_summary,
    solver_results: dict,
) -> bytes:
    """Multi-sheet workbook with config, valuations and risk metrics."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:

        # Sheet 1 — Config (flattened)
        cfg_dict = asdict(config)
        cfg_rows = []
        for section, vals in cfg_dict.items():
            for k, v in vals.items():
                if isinstance(v, list):
                    cfg_rows.append((section, k, ", ".join(f"{x}" for x in v)))
                else:
                    cfg_rows.append((section, k, v))
        pd.DataFrame(cfg_rows, columns=["Section", "Parameter", "Value"]).to_excel(
            writer, sheet_name="Config", index=False
        )

        # Sheet 2 — Central scenario
        mc_result.central_df.to_excel(writer, sheet_name="Central Scenario", index=False)

        # Sheet 3 — Risk summary
        risk_rows = []
        for c in risk_summary.combinations:
            risk_rows.append({
                "Supply": c.supply_structure,
                "Pricing": c.pricing_structure,
                "P10 (€M)": c.joint.p10 / 1e6,
                "P50 (€M)": c.joint.p50 / 1e6,
                "P90 (€M)": c.joint.p90 / 1e6,
                "VaR-95 (€M)": c.joint.var_95 / 1e6,
                "ES-95 (€M)": c.joint.es_95 / 1e6,
                "P(NPV<0)": c.joint.prob_negative,
                "Price share": c.variance_decomp.price_share,
                "Volume share": c.variance_decomp.volume_share,
                "Interaction": c.variance_decomp.interaction_share,
            })
        pd.DataFrame(risk_rows).to_excel(writer, sheet_name="Risk Summary", index=False)

        # Sheet 4 — NPV Paths (joint mode pivoted)
        paths_df = mc_result.paths_df
        joint = paths_df[paths_df["mode"] == "joint"].copy()
        joint["combo"] = joint["supply_structure"] + " / " + joint["pricing_structure"]
        try:
            pivot = joint.pivot_table(
                index="path", columns="combo", values="producer_npv"
            )
            pivot.to_excel(writer, sheet_name="NPV Paths")
        except Exception:
            joint.to_excel(writer, sheet_name="NPV Paths", index=False)

        # Sheet 5 — Break-even
        neg = solver_results.get("negotiation", {})
        neg_rows = []
        for supply_name, n in neg.items():
            neg_rows.append({
                "Supply": supply_name,
                "Producer floor (€/MWh)": n.producer_floor.strike_eur_mwh,
                "Consumer ceiling (€/MWh)": n.consumer_ceiling.strike_eur_mwh,
                "Midpoint (€/MWh)": n.midpoint,
                "Spread (€/MWh)": n.spread_eur_mwh,
            })
        pd.DataFrame(neg_rows).to_excel(writer, sheet_name="Break-even", index=False)

        # Sheet 6 — Tornado
        tornado = solver_results.get("tornado_pap_fixedflat")
        if tornado is not None:
            t_rows = []
            for e in tornado.entries:
                t_rows.append({
                    "Parameter": e.parameter,
                    "Base value": e.base_value,
                    "Low value": e.low_value,
                    "High value": e.high_value,
                    "NPV base (€M)": e.npv_base / 1e6,
                    "NPV low (€M)": e.npv_low / 1e6,
                    "NPV high (€M)": e.npv_high / 1e6,
                    "Δ low (€M)": e.delta_npv_low / 1e6,
                    "Δ high (€M)": e.delta_npv_high / 1e6,
                    "Abs swing (€M)": e.abs_swing / 1e6,
                })
            pd.DataFrame(t_rows).to_excel(writer, sheet_name="Tornado", index=False)

        # Sheet 7 — Variance decomposition
        var_rows = []
        for c in risk_summary.combinations:
            var_rows.append({
                "Supply": c.supply_structure,
                "Pricing": c.pricing_structure,
                "Total variance": c.variance_decomp.total_variance,
                "Price share": c.variance_decomp.price_share,
                "Volume share": c.variance_decomp.volume_share,
                "Interaction share": c.variance_decomp.interaction_share,
            })
        pd.DataFrame(var_rows).to_excel(writer, sheet_name="Variance Decomp", index=False)

        # ---- Light conditional formatting on key sheets ----
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        green = PatternFill("solid", fgColor="DCFCE7")
        red = PatternFill("solid", fgColor="FEE2E2")

        ws = writer.book["Risk Summary"]
        # Columns C..G are the NPV columns (1-indexed after Supply/Pricing)
        for col in ("C", "D", "E", "F", "G"):
            rng = f"{col}2:{col}{ws.max_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=red))
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=green))

    return bio.getvalue()
